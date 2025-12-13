# ensemble_forecast.py
# Week 8: Complete Quantitative AI Forecasting System (Part 2)

import warnings
warnings.filterwarnings("ignore")

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

from forecast_system import VolumeForecaster
from ml_forecast import MLForecaster


class EnsembleForecaster:
    """
    Combines Week 7 statistical methods + Week 8 ML methods into a robust ensemble,
    generates 13-month forecast, creates Excel deliverable, and saves visualization suite.
    """

    def __init__(self, data_path: str):
        print("\033[91m" + "=" * 70 + "\033[0m")
        print("\033[97m" + "🇺🇸 ENSEMBLE AI FORECASTING SYSTEM (Week 8)" + "\033[0m")
        print("\033[94m" + "=" * 70 + "\033[0m")

        # Load data
        if not os.path.exists(data_path):
            raise FileNotFoundError(f"Data not found at: {data_path}")

        self.data_path = data_path
        self.data = pd.read_csv(data_path)

        # normalize date index
        if "date" in self.data.columns:
            self.data["date"] = pd.to_datetime(self.data["date"])
            self.data = self.data.set_index("date")

        # component systems
        self.stat_forecaster = VolumeForecaster(data_path)   # Week 7
        self.ml_forecaster = MLForecaster(self.data)         # Week 8

        # weights (from your screenshots)
        self.weights = {
            "moving_average": 0.15,
            "exponential_smoothing": 0.15,
            "linear_regression": 0.20,
            "random_forest": 0.25,
            "xgboost": 0.25,   # if xgb not available, we’ll renormalize
        }

    def create_ensemble_forecast(self, periods_days: int = 13 * 30):
        print("\n\033[96m⚙ CREATING ENSEMBLE FORECAST...\033[0m")

        # Week 7 statistical forecasts (next-period only, but we can use them to adjust early horizon)
        ma_next, _ = self.stat_forecaster.moving_average_forecast(window=7)
        exp_next, _ = self.stat_forecaster.exponential_smoothing(alpha=0.3)

        # Train ML models and create daily forecast horizon
        self.ml_forecaster.train_models(test_size=90, feature_window=30)
        ml_daily = self.ml_forecaster.forecast_future(periods=periods_days, feature_window=30)

        # ML column mapping
        has_xgb = "xgb" in ml_daily.columns

        # Renormalize weights if xgb not available
        weights = self.weights.copy()
        if not has_xgb:
            weights.pop("xgboost")
            s = sum(weights.values())
            weights = {k: v / s for k, v in weights.items()}

        # Build weighted ML ensemble (daily)
        # Map names
        linear_col = "linear"
        rf_col = "rf"
        xgb_col = "xgb" if has_xgb else None

        ml_daily["weighted_ensemble"] = (
            weights["linear_regression"] * ml_daily[linear_col] +
            weights["random_forest"] * ml_daily[rf_col] +
            (weights["xgboost"] * ml_daily[xgb_col] if has_xgb else 0.0)
        )

        # Apply statistical adjustment to first 30 days (as in screenshots)
        stat_adjustment = (weights["moving_average"] * ma_next) + (weights["exponential_smoothing"] * exp_next)
        ml_daily.loc[:29, "weighted_ensemble"] = (
            0.7 * ml_daily.loc[:29, "weighted_ensemble"] + 0.3 * stat_adjustment
        )

        # Confidence interval (std of model preds)
        model_cols = [linear_col, rf_col] + ([xgb_col] if has_xgb else [])
        std = ml_daily[model_cols].std(axis=1)
        ml_daily["lower_bound"] = ml_daily["weighted_ensemble"] - 1.96 * std
        ml_daily["upper_bound"] = ml_daily["weighted_ensemble"] + 1.96 * std

        # non-negative
        ml_daily["lower_bound"] = ml_daily["lower_bound"].clip(lower=0)

        print("✓ Created weighted ensemble forecast")
        print(f"✓ Weights used: {weights}")

        return ml_daily

    def create_excel_workbook(self, forecast_df: pd.DataFrame, out_path: str = "volume_forecast_13months.xlsx"):
        print("\n\033[94m📊 CREATING EXCEL WORKBOOK...\033[0m")

        wb = Workbook()

        # Patriotic fills
        red_fill = PatternFill(start_color="B22234", end_color="B22234", fill_type="solid")
        blue_fill = PatternFill(start_color="3C3B6E", end_color="3C3B6E", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="3C3B6E")

        thin = Side(style="thin")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # =========================
        # Sheet 1: Dashboard
        # =========================
        ws1 = wb.active
        ws1.title = "📌 Dashboard"

        ws1["B2"] = "AI VOLUME FORECASTING SYSTEM"
        ws1["B2"].font = title_font
        ws1.merge_cells("B2:H2")

        ws1["B4"] = "13-Month Forecast Summary"
        ws1["B4"].font = Font(bold=True, size=14)

        # Monthly aggregation for summary table
        tmp = forecast_df.copy()
        tmp["date"] = pd.to_datetime(tmp["date"])
        tmp = tmp.set_index("date")

        monthly = tmp.resample("M").agg({
            "weighted_ensemble": "sum",
            "lower_bound": "sum",
            "upper_bound": "sum"
        })

        headers = ["Month", "Forecast", "Lower Bound", "Upper Bound", "Confidence"]
        for col, h in enumerate(headers, start=2):
            cell = ws1.cell(row=6, column=col, value=h)
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for r, (idx, row) in enumerate(monthly.iterrows(), start=7):
            ws1.cell(row=r, column=2, value=idx.strftime("%b %Y"))
            ws1.cell(row=r, column=3, value=float(row["weighted_ensemble"]))
            ws1.cell(row=r, column=4, value=float(row["lower_bound"]))
            ws1.cell(row=r, column=5, value=float(row["upper_bound"]))
            ws1.cell(row=r, column=6, value="95%")

            # alternating stripes
            fill = red_fill if (r % 2 == 0) else white_fill
            for c in range(2, 7):
                cell = ws1.cell(row=r, column=c)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if fill == red_fill:
                    cell.font = Font(color="FFFFFF")

        # Chart
        chart = LineChart()
        chart.title = "13-Month Volume Forecast"
        chart.y_axis.title = "Volume"
        chart.x_axis.title = "Month"
        chart.style = 13

        data = Reference(ws1, min_col=3, min_row=6, max_row=6 + len(monthly), max_col=5)
        cats = Reference(ws1, min_col=2, min_row=7, max_row=6 + len(monthly))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.width = 18
        chart.height = 10
        ws1.add_chart(chart, "I6")

        # =========================
        # Sheet 2: Daily Forecasts
        # =========================
        ws2 = wb.create_sheet("📅 Daily Forecasts")

        daily_headers = ["Date", "Linear", "Random Forest", "XGBoost", "Ensemble", "Lower", "Upper"]
        for col, h in enumerate(daily_headers, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = blue_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        has_xgb = "xgb" in forecast_df.columns

        for r, row in enumerate(forecast_df.itertuples(index=False), start=2):
            # row fields depend on df columns; access by name via dict
            d = row._asdict()
            ws2.cell(row=r, column=1, value=str(d["date"])[:10])
            ws2.cell(row=r, column=2, value=float(d.get("linear", 0)))
            ws2.cell(row=r, column=3, value=float(d.get("rf", 0)))
            ws2.cell(row=r, column=4, value=float(d.get("xgb", 0)) if has_xgb else "")
            ws2.cell(row=r, column=5, value=float(d["weighted_ensemble"]))
            ws2.cell(row=r, column=6, value=float(d["lower_bound"]))
            ws2.cell(row=r, column=7, value=float(d["upper_bound"]))

            for c in range(1, 8):
                ws2.cell(row=r, column=c).border = thin_border

        # =========================
        # Sheet 3: Model Performance
        # =========================
        ws3 = wb.create_sheet("📈 Model Performance")
        ws3["B2"] = "Model Performance Comparison"
        ws3["B2"].font = Font(bold=True, size=14)

        perf_headers = ["Model", "MAE", "RMSE", "R²"]
        for col, h in enumerate(perf_headers, start=2):
            cell = ws3.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = blue_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # pull real metrics from MLForecaster
        metrics = self.ml_forecaster.metrics
        rows = []
        if "linear" in metrics:
            rows.append(("Linear Regression", metrics["linear"]["mae"], metrics["linear"]["rmse"], metrics["linear"]["r2"]))
        if "rf" in metrics:
            rows.append(("Random Forest", metrics["rf"]["mae"], metrics["rf"]["rmse"], metrics["rf"]["r2"]))
        if "xgb" in metrics:
            rows.append(("XGBoost", metrics["xgb"]["mae"], metrics["xgb"]["rmse"], metrics["xgb"]["r2"]))

        for r, (name, mae, rmse, r2) in enumerate(rows, start=5):
            ws3.cell(row=r, column=2, value=name)
            ws3.cell(row=r, column=3, value=float(mae))
            ws3.cell(row=r, column=4, value=float(rmse))
            ws3.cell(row=r, column=5, value=float(r2))
            for c in range(2, 6):
                ws3.cell(row=r, column=c).border = thin_border

        wb.save(out_path)
        print(f"✓ Excel workbook created: {out_path}")
        return out_path

    def create_visualization_suite(self, forecast_df: pd.DataFrame, out_path: str = "ensemble_forecast_visuals.png"):
        print("\n\033[92m🖼 CREATING VISUALIZATION SUITE...\033[0m")

        df = forecast_df.copy()
        df["date"] = pd.to_datetime(df["date"])

        # last 180 historical for context
        hist = self.data.copy().reset_index()
        hist = hist.tail(180)

        fig, axes = plt.subplots(2, 2, figsize=(15, 10))
        fig.suptitle("AI Volume Forecasting System - 13 Month Outlook", fontsize=16, fontweight="bold")

        # Plot 1: Historical + Forecast + CI
        axes[0, 0].plot(hist["date"], hist["volume"], label="Historical", linewidth=2)
        axes[0, 0].plot(df["date"], df["weighted_ensemble"], label="Forecast", linewidth=2)
        axes[0, 0].fill_between(df["date"], df["lower_bound"], df["upper_bound"], alpha=0.3, label="95% CI")
        axes[0, 0].set_title("Historical Data & 13-Month Forecast")
        axes[0, 0].set_xlabel("Date")
        axes[0, 0].set_ylabel("Volume")
        axes[0, 0].legend()
        axes[0, 0].grid(True, alpha=0.3)

        # Plot 2: Model comparison (monthly mean)
        m = df.set_index("date").resample("M").mean()
        x = np.arange(len(m))
        width = 0.25
        axes[0, 1].bar(x - width, m.get("linear", 0), width, label="Linear")
        axes[0, 1].bar(x, m.get("rf", 0), width, label="Random Forest")
        if "xgb" in m.columns:
            axes[0, 1].bar(x + width, m["xgb"], width, label="XGBoost")
        axes[0, 1].set_title("Model Predictions by Month")
        axes[0, 1].set_xlabel("Month")
        axes[0, 1].set_ylabel("Average Daily Volume")
        axes[0, 1].legend()
        axes[0, 1].grid(True, alpha=0.3, axis="y")

        # Plot 3: Monthly totals (ensemble)
        monthly_totals = df.set_index("date").resample("M").sum()
        axes[1, 0].bar(range(len(monthly_totals)), monthly_totals["weighted_ensemble"])
        axes[1, 0].set_title("Monthly Volume Totals (13-Month Forecast)")
        axes[1, 0].set_xlabel("Month")
        axes[1, 0].set_ylabel("Total Volume")
        axes[1, 0].set_xticks(range(len(monthly_totals)))
        axes[1, 0].set_xticklabels([d.strftime("%b %y") for d in monthly_totals.index], rotation=45)
        axes[1, 0].grid(True, alpha=0.3, axis="y")

        # Plot 4: Uncertainty %
        uncertainty = (df["upper_bound"] - df["lower_bound"]) / df["weighted_ensemble"].replace(0, np.nan) * 100
        axes[1, 1].plot(df["date"], uncertainty)
        axes[1, 1].fill_between(df["date"], uncertainty, alpha=0.3)
        axes[1, 1].set_title("Forecast Uncertainty Over Time")
        axes[1, 1].set_xlabel("Date")
        axes[1, 1].set_ylabel("Uncertainty (%)")
        axes[1, 1].grid(True, alpha=0.3)

        plt.tight_layout()
        plt.savefig(out_path, dpi=150, bbox_inches="tight")
        print(f"✓ Visualization suite saved: {out_path}")
        return out_path


def _resolve_data_path():
    # supports either data/historical_volumes.csv or historical_volumes.csv
    if os.path.exists("data/historical_volumes.csv"):
        return "data/historical_volumes.csv"
    if os.path.exists("historical_volumes.csv"):
        return "historical_volumes.csv"
    raise FileNotFoundError("Could not find historical_volumes.csv in ./data or repo root.")


def main():
    print("\033[91m" + "=" * 70 + "\033[0m")
    print("\033[97m" + "🇺🇸 COMPLETE AI FORECASTING SYSTEM (Week 8)" + "\033[0m")
    print("\033[94m" + "=" * 70 + "\033[0m")

    data_path = _resolve_data_path()

    ensemble = EnsembleForecaster(data_path)

    forecast_df = ensemble.create_ensemble_forecast(periods_days=13 * 30)

    excel_path = ensemble.create_excel_workbook(forecast_df, out_path="volume_forecast_13months.xlsx")
    visuals_path = ensemble.create_visualization_suite(forecast_df, out_path="ensemble_forecast_visuals.png")

    # Final summary
    tmp = forecast_df.copy()
    tmp["date"] = pd.to_datetime(tmp["date"])
    tmp = tmp.set_index("date")
    monthly = tmp.resample("M").agg({"weighted_ensemble": "sum", "lower_bound": "sum", "upper_bound": "sum"})

    print("\n📌 13-MONTH FORECAST SUMMARY:")
    print(f" Total Volume Forecast: {monthly['weighted_ensemble'].sum():.0f} units")
    print(f" Avg Monthly Volume:    {monthly['weighted_ensemble'].mean():.0f} units/month")
    print(f" Peak Month:            {monthly['weighted_ensemble'].idxmax().strftime('%b %Y')}")
    print(f" Peak Volume:           {monthly['weighted_ensemble'].max():.0f} units")

    print("\n✅ DELIVERABLES CREATED:")
    print(f" - {excel_path} (Excel workbook)")
    print(f" - {visuals_path} (Visualization suite)")
    print(" - Full codebase in your GitHub repo")

    return forecast_df


if __name__ == "__main__":
    main()